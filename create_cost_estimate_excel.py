#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to generate simplified Cost Estimate Excel - Summary only
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_summary_cost_excel():
    """Create summary cost estimate Excel - main modules only"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Chi Phí Theo Module"
    
    # Define styles
    header_fill = PatternFill(start_color="556B2F", end_color="556B2F", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    total_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    total_font = Font(name="Arial", size=11, bold=True)
    
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_alignment = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    row = 1
    
    # Title
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "BẢNG CHI PHÍ DEMO - CRM_DQP10"
    cell.font = Font(name="Arial", size=14, bold=True, color="556B2F")
    cell.alignment = center_alignment
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = "Hệ Thống Quản Lý Dân Quân Phường 10"
    cell.font = Font(name="Arial", size=11, italic=True)
    cell.alignment = center_alignment
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = f"Ngày: {datetime.now().strftime('%d/%m/%Y')}"
    cell.font = Font(name="Arial", size=10)
    cell.alignment = center_alignment
    row += 2
    
    # Table Header
    headers = ['STT', 'Module', 'Thiết kế UI\n(VNĐ)', 'Phân tích\nNghiệp vụ\n(VNĐ)', 'Tài liệu\nHướng dẫn\n(VNĐ)', 'Tổng\n(VNĐ)']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    row += 1
    
    # Data - Summary only (no sub-items)
    modules_data = [
        ['1', 'Module 1: Quản lý Hoạt động (Activities)', '280,000', '400,000', '180,000', '860,000'],
        ['2', 'Module 2: Quản lý Nhân sự (Personnel)', '120,000', '350,000', '30,000', '500,000'],
        ['3', 'Module 3: Quản lý Thiết bị (Inventory)', '100,000', '300,000', '30,000', '430,000'],
        ['4', 'Module 4: Quản lý Văn bản (Documents)', '80,000', '250,000', '30,000', '360,000'],
        ['5', 'Module 5: Quản lý Phân quyền (Permissions)', '80,000', '250,000', '30,000', '360,000'],
        ['6', 'Module 6: Dashboard & Báo cáo', '120,000', '250,000', '30,000', '400,000'],
        ['7', 'Settings & Planning (Cài đặt)', '240,000', '200,000', '70,000', '510,000'],
        ['8', 'Dashboard Planning (Báo cáo kế hoạch)', '160,000', '-', '-', '160,000'],
        ['9', 'Landing Page & Design System', '620,000', '-', '600,000', '1,220,000'],
    ]
    
    for data_row in modules_data:
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.border = thin_border
            
            if col_num == 1:  # STT
                cell.font = bold_font
                cell.alignment = center_alignment
            elif col_num == 2:  # Module name
                cell.font = bold_font
                cell.alignment = left_alignment
            else:  # Numbers
                cell.font = normal_font
                cell.alignment = right_alignment
                
        row += 1
    
    # Total row
    cell = ws[f'A{row}']
    cell.value = ''
    cell.border = thin_border
    cell.fill = total_fill
    
    ws.merge_cells(f'B{row}:B{row}')
    cell = ws[f'B{row}']
    cell.value = 'TỔNG CỘNG'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = right_alignment
    cell.border = thin_border
    
    cell = ws[f'C{row}']
    cell.value = '2,000,000'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = right_alignment
    cell.border = thin_border
    
    cell = ws[f'D{row}']
    cell.value = '2,000,000'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = right_alignment
    cell.border = thin_border
    
    cell = ws[f'E{row}']
    cell.value = '1,000,000'
    cell.font = total_font
    cell.fill = total_fill
    cell.alignment = right_alignment
    cell.border = thin_border
    
    cell = ws[f'F{row}']
    cell.value = '5,000,000'
    cell.font = Font(name="Arial", size=12, bold=True, color="FF0000")
    cell.fill = total_fill
    cell.alignment = right_alignment
    cell.border = thin_border
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = '(Năm triệu đồng)'
    cell.font = Font(name="Arial", size=10, italic=True)
    cell.alignment = center_alignment
    row += 2
    
    # Notes
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = 'GHI CHÚ:'
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = left_alignment
    row += 1
    
    notes = [
        '• Thiết kế UI: Bao gồm HTML prototypes, responsive design, và UI components',
        '• Phân tích Nghiệp vụ: User flows, Data flows, Business rules cho từng module',
        '• Tài liệu Hướng dẫn: User guides, Technical docs, Design system documentation',
        '• Tổng số files: 27 files (14 HTML prototypes + 4 Design System + 5 Technical Docs + 4 Others)',
        '• Timeline: 18 ngày làm việc',
    ]
    
    for note in notes:
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = note
        cell.font = Font(name="Arial", size=9)
        cell.alignment = left_alignment
        row += 1
    
    # Save file
    filename = f"CRM_DQP10_Chi_Phi_Theo_Module_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = f"c:\\Users\\VietLong\\Downloads\\Github\\CRM_DQP10\\{filename}"
    wb.save(filepath)
    
    return filepath

if __name__ == "__main__":
    try:
        filepath = create_summary_cost_excel()
        print(f"✅ Đã tạo file Excel thành công!")
        print(f"📁 Đường dẫn: {filepath}")
    except Exception as e:
        print(f"❌ Lỗi: {e}")
